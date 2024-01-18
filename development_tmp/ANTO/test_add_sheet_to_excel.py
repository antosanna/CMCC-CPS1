import openpyxl
import os
import pandas as pd

import openpyxl as xl

path1 = 'test_sps4_hindcast_IC_NEMO_list.zeus.xlsx'
path2 = 'SPS4-hindcast-production-list.xlsx'

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path2)
