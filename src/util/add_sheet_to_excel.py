import openpyxl as xl
import pandas as pd
import argparse
import os

def argParser():
    """
    Argument parser
    """
    parser = argparse.ArgumentParser(prog='add_sheet_to_excel',description='Add a sheet to an existing workbook xls.')
    parser.add_argument("orig_sheet", help="sheet_to_add")
    parser.add_argument("dst_workbook", help="Destination workbook")
    parser.add_argument("-p", "--path", default=".",
            help='file path (default current directory)')
#    parser.add_argument("-l", "--logdir", default=".",
#            help='log file path (default current directory)')
    args = parser.parse_args()
    return(args)

def check_path_exists(path):
    """
    Check if path exists
    """
    if not (os.path.exists(path)):
        raise InputError('[INPUTERROR] Unknown path')

def myaddworkbook():
    # File path
   path1 = os.path.join(args.path, args.orig_sheet)
   path2 = os.path.join(args.path, args.dst_workbook)
   check_path_exists(path2)
   check_path_exists(path1)

   wb1 = xl.load_workbook(filename=path1)
   ws1 = wb1.worksheets[0]

   wb2 = xl.load_workbook(filename=path2)
   ws2 = wb2.create_sheet(ws1.title)

   for row in ws1:
       for cell in row:
           ws2[cell.coordinate].value = cell.value

   wb2.save(path2)

if __name__ == '__main__':
   args = argParser()
   myaddworkbook()
